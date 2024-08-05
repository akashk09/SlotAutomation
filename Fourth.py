import openpyxl
from datetime import datetime, time
import os
import re

def validate_excel_file(file_path):
    #checks if the file exists at the given path
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Read the Excel file, starting from row 2 (index 1 in pandas)
    if sheet.max_row > 6:
        print("Error: The sheet has more than 5 data rows.")
        return
    
    # Check for mandatory columns
    mandatory_columns = {
        'MIPS': float,
        'date': (str, datetime),
        'start time': (str, time, datetime),
        'end time': (str, time, datetime)
    }

    column_indices = {}
    for col in sheet[1]:
        if col.value in mandatory_columns:
            column_indices[col.value] = col.column

    if len(column_indices) != len(mandatory_columns):
        print("Error: Not all mandatory columns are present in the sheet.")
        return

    errors_found = False

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        for col_name, col_index in column_indices.items():
            value = row[col_index - 1]
            
            if value is None:
                print(f"Error in row {row_index}: {col_name} is empty.")
                errors_found = True
                continue

            if col_name == 'MIPS':
                try:
                    float(value)
                except ValueError:
                    print(f"Error in row {row_index}: MIPS '{value}' is not a valid number.")
                    errors_found = True

            elif col_name == 'date':
                if isinstance(value, datetime):
                    pass
                elif isinstance(value, str):
                    try:
                        datetime.strptime(value, '%Y-%m-%d')
                    except ValueError:
                        print(f"Error in row {row_index}: Invalid date format '{value}'. Use YYYY-MM-DD.")
                        errors_found = True
                else:
                    print(f"Error in row {row_index}: Date '{value}' is neither a string nor a datetime object.")
                    errors_found = True

            elif col_name in ['start time', 'end time']:
                if isinstance(value, (datetime, time)):
                    pass
                elif isinstance(value, str):
                    # Remove any whitespace and 'BST' if present
                    cleaned_value = value.strip().upper().replace('BST', '').strip()
                    # Use regex to check if the time is in valid formats
                    if re.match(r'^(\d{1,2}(:\d{2})?|\d{3,4})$', cleaned_value):
                        try:
                            if ':' in cleaned_value:
                                datetime.strptime(cleaned_value, '%H:%M')
                            else:
                                if len(cleaned_value) <= 2:
                                    datetime.strptime(cleaned_value, '%H')
                                else:
                                    hours = int(cleaned_value[:-2])
                                    minutes = int(cleaned_value[-2:])
                                    if 0 <= hours <= 23 and 0 <= minutes <= 59:
                                        time(hours, minutes)
                                    else:
                                        raise ValueError
                        except ValueError:
                            print(f"Error in row {row_index}: Invalid time format '{value}' for {col_name}. Use HH, HH:MM, or HHMM (24-hour format), optionally followed by BST.")
                            errors_found = True
                    else:
                        print(f"Error in row {row_index}: Invalid time format '{value}' for {col_name}. Use HH, HH:MM, or HHMM (24-hour format), optionally followed by BST.")
                        errors_found = True
                else:
                    print(f"Error in row {row_index}: {col_name} '{value}' is not a valid time format.")
                    errors_found = True

    if not errors_found:
        print("Validation complete. No errors found.")
    else:
        print("Validation complete. Errors were found.")

# Specify the file path
file_path = r"D:\Akash Kumar\JOB WORK\Resume\Fly\Project X\SlotAttachment.xlsx"

# Run the validation
validate_excel_file(file_path)
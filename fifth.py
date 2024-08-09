import openpyxl
from datetime import datetime, time
import os
import re
from jira import JIRA
import tempfile

def validate_excel_file(file_path, jira, issue):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    if sheet.max_row > 6:
        comment = "Error: The sheet has more than 5 data rows."
        jira.add_comment(issue, comment)
        print(comment)
        return

    mandatory_columns = {
        'MIPS': float,
        'date': (str, datetime),
        'start time': (str, time, datetime),
        'end time': (str, time, datetime)
    }

    column_indices = {}
    for col in sheet[1]:
        if col.value in mandatory_columns:
            column_indices[col.value] = col.column

    if len(column_indices) != len(mandatory_columns):
        comment = "Error: Not all mandatory columns are present in the sheet."
        jira.add_comment(issue, comment)
        print(comment)
        return

    errors_found = False

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        for col_name, col_index in column_indices.items():
            value = row[col_index - 1]
            
            if value is None:
                comment = f"Error in row {row_index}: {col_name} is empty."
                jira.add_comment(issue, comment)
                print(comment)
                errors_found = True
                continue

            if col_name == 'MIPS':
                try:
                    float(value)
                except ValueError:
                    comment = f"Error in row {row_index}: MIPS '{value}' is not a valid number."
                    jira.add_comment(issue, comment)
                    print(comment)
                    errors_found = True

            elif col_name == 'date':
                if isinstance(value, datetime):
                    pass
                elif isinstance(value, str):
                    try:
                        datetime.strptime(value, '%Y-%m-%d')
                    except ValueError:
                        comment = f"Error in row {row_index}: Invalid date format '{value}'. Use YYYY-MM-DD."
                        jira.add_comment(issue, comment)
                        print(comment)
                        errors_found = True
                else:
                    comment = f"Error in row {row_index}: Date '{value}' is neither a string nor a datetime object."
                    jira.add_comment(issue, comment)
                    print(comment)
                    errors_found = True

            elif col_name in ['start time', 'end time']:
                if isinstance(value, (datetime, time)):
                    pass
                elif isinstance(value, str):
                    cleaned_value = value.strip().upper().replace('BST', '').strip()
                    if re.match(r'^(\d{1,2}(:\d{2})?|\d{3,4})$', cleaned_value):
                        try:
                            if ':' in cleaned_value:
                                datetime.strptime(cleaned_value, '%H:%M')
                            else:
                                if len(cleaned_value) <= 2:
                                    datetime.strptime(cleaned_value, '%H')
                                else:
                                    hours = int(cleaned_value[:-2])
                                    minutes = int(cleaned_value[-2:])
                                    if 0 <= hours <= 23 and 0 <= minutes <= 59:
                                        time(hours, minutes)
                                    else:
                                        raise ValueError
                        except ValueError:
                            comment = f"Error in row {row_index}: Invalid time format '{value}' for {col_name}. Use HH, HH:MM, or HHMM (24-hour format), optionally followed by BST."
                            jira.add_comment(issue, comment)
                            print(comment)
                            errors_found = True
                    else:
                        comment = f"Error in row {row_index}: Invalid time format '{value}' for {col_name}. Use HH, HH:MM, or HHMM (24-hour format), optionally followed by BST."
                        jira.add_comment(issue, comment)
                        print(comment)
                        errors_found = True
                else:
                    comment = f"Error in row {row_index}: {col_name} '{value}' is not a valid time format."
                    jira.add_comment(issue, comment)
                    print(comment)
                    errors_found = True

    if not errors_found:
        comment = "Validation complete. No errors found."
        jira.add_comment(issue, comment)
        print(comment)
    else:
        comment = "Validation complete. Errors were found. Please check the comments above for details."
        jira.add_comment(issue, comment)
        print(comment)

def process_jira_ticket(jira, issue_key):
    issue = jira.issue(issue_key)
    
    # Find the Excel attachment
    excel_attachment = None
    for attachment in issue.fields.attachment:
        if attachment.filename.endswith('.xlsx'):
            excel_attachment = attachment
            break
    
    if not excel_attachment:
        comment = "Error: No Excel file found in the ticket attachments."
        jira.add_comment(issue, comment)
        print(comment)
        return
    
    # Download the attachment to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        temp_file.write(excel_attachment.get())
        temp_file_path = temp_file.name
    
    try:
        validate_excel_file(temp_file_path, jira, issue)
    finally:
        # Clean up the temporary file
        os.unlink(temp_file_path)

# JIRA connection details
jira_options = {'server': 'https://your-jira-instance.com'}
jira = JIRA(options=jira_options, basic_auth=('your_username', 'your_password'))

# Specify the JIRA ticket key
issue_key = 'PROJECT-123'

# Process the JIRA ticket
process_jira_ticket(jira, issue_key)